import streamlit as st
import pandas as pd
import gspread
import zipfile
import re
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from openpyxl import load_workbook

# Scope
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

# Credenciales
creds = ServiceAccountCredentials.from_json_keyfile_name(r'C:\Users\OAM-GAL-061\Documents\Python\credentials.json', scope)
client = gspread.authorize(creds)

# Hoja de cálculo
spreadsheet = client.open("BANCOS PY-GS")

meses = {
    'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04',
    'may': '05', 'jun': '06', 'jul': '07', 'ago': '08',
    'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'
}

def convertir_fecha(fecha):
    dia, mes, anio = fecha.split('-')
    mes = meses[mes]
    return f"{dia}/{mes}/{anio}"

# App
opcion = st.sidebar.selectbox("Menú", ["Pagos", "Lecturas"])

if opcion == "Pagos":
    st.title("Actualización de pagos y descarga de bancos - MAYO")
    # Subir archivo
    fecha_corte = st.date_input("Selecciona la fecha de corte")
    uploaded_file = st.file_uploader("Elige tu tabla de cobro", type=["xlsx", "xls"])
    if uploaded_file is not None:
        if st.button("Procesar"):
            msjcobro_proceso = st.empty()
            msjcobro_proceso.write("Procesando archivo...")
            ########################################### IPQ ############################################
            ###############  2239  ###############
            sheet2239 = spreadsheet.worksheet("2239")
            # Obtener datos
            data2239 = sheet2239.get_all_records()

            # DataFrame y Limpieza
            df2239_T= pd.DataFrame(data2239)
            df2239 = df2239_T.drop_duplicates() # Eliminar duplicados
            df2239 = df2239.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df2239.columns = df2239.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df2239['OTROS'].str.contains('Consumo') # Filtro de Consumo
            condSancion = df2239['OTROS'].str.contains('Sancion') # Filtro de Sancion
            condMedidor = df2239['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df2239['OTROS'].str.contains('Reconexion') # Filtro de Reconexion
            condGasto = df2239['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df2239['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df2239['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df2239['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df2239['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df2239['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df2239['OTROS'].str.contains('Caja de medidor') # Filtro de Caja de medidor
            condConv = df2239['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df2239['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df2239 = df2239[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df2239 = df2239.filter(items = ['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df2239['INGRESOS'] = df2239['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df2239['FECHA'] = pd.to_datetime(df2239['FECHA'], format='%d/%m/%Y')
            df2239['FECHA'] = df2239['FECHA'].dt.strftime('%d/%m/%Y')
            df2239['OTROS'] = df2239['OTROS'].str.upper()

            #df2239 = df2239.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df2239 = df2239.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df2239 = df2239[['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df2239.insert(0,'CUENTA', "BAJIO 2239")

            #print(df2239)

            ###############  2162  ############### 

            sheet2162 = spreadsheet.worksheet("2162")

            # Obtener datos
            data2162 = sheet2162.get_all_values()

            # DataFrame y Limpieza
            columnas2162 = ['FECHA','DESCRIPCION','RETIROS','INGRESOS','TOTALES', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS', 'DIFERENCIA','RECIBO', 'POLIZA IC', 'POLIZA DC', '#1', '#2', '#3', '#4', '#5', '#6', '#7','#8','#9','#10','#11','#12','#13','#14','#15','#16','#17']
            df2162_T = pd.DataFrame(data2162[1:], columns = columnas2162)
            df2162 = df2162_T.drop_duplicates() # Eliminar duplicados
            df2162 = df2162.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios

            condConsumo = df2162['OTROS'].str.contains('Consumo') # Filtro de Consumo
            condSancion = df2162['OTROS'].str.contains('Sancion') # Filtro de Sancion
            condMedidor = df2162['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df2162['OTROS'].str.contains('Reconexion') # Filtro de Reconexion
            condGasto = df2162['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df2162['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df2162['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df2162['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df2162['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df2162['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df2162['OTROS'].str.contains('Caja de medidor') # Filtro de Caja de medidor
            condConv = df2162['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df2162['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df2162 = df2162[condConsumo | condSancion  | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df2162 = df2162.filter(items = ['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'TOTALES','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df2162['INGRESOS'] = df2162['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df2162['FECHA'] = pd.to_datetime(df2162['FECHA'], format='%d/%m/%Y')
            df2162['FECHA'] = df2162['FECHA'].dt.strftime('%d/%m/%Y')
            df2162['OTROS'] = df2162['OTROS'].str.upper()

            #df2162 = df2162.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df2162 = df2162.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df2162 = df2162[['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df2162.insert(0,'CUENTA', "BAJIO 2162")

            #print(df2162)

            ########################################### FRO ############################################

            ###############  9721  ############### 

            sheet9721 = spreadsheet.worksheet("9721")

            # Obtener datos
            data9721 = sheet9721.get_all_values()

            # DataFrame y Limpieza
            columnas9721 = ['FECHA','DESCRIPCION','RETIROS','INGRESOS','TOTALES', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS', 'DIFERENCIA','RECIBO', 'POLIZA IC', 'POLIZA DC']
            df9721_T = pd.DataFrame(data9721[1:], columns = columnas9721)
            df9721 = df9721_T.drop_duplicates() # Eliminar duplicados
            df9721 = df9721.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios

            condConsumo = df9721['OTROS'].str.contains('Consumo') # Filtro de Consumo
            condSancion = df9721['OTROS'].str.contains('Sancion') # Filtro de Sancion
            condMedidor = df9721['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df9721['OTROS'].str.contains('Reconexion') # Filtro de Reconexion
            condGasto = df9721['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df9721['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df9721['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df9721['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df9721['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df9721['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df9721['OTROS'].str.contains('Caja de medidor') # Filtro de Caja de medidor
            condConv = df9721['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df9721['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df9721 = df9721[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df9721 = df9721.filter(items = ['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'TOTALES','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df9721['INGRESOS'] = df9721['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df9721['FECHA'] = pd.to_datetime(df9721['FECHA'], format='%d/%m/%Y')
            df9721['FECHA'] = df9721['FECHA'].dt.strftime('%d/%m/%Y')
            df9721['OTROS'] = df9721['OTROS'].str.upper()

            #df9721 = df9721.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df9721 = df9721.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df9721 = df9721[['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df9721.insert(0,'CUENTA', "BAJIO 9721")

            #print(df9721)

            ###############  234  ############### 

            sheet234 = spreadsheet.worksheet("234")

            # Obtener datos
            data234 = sheet234.get_all_records()

            # DataFrame y Limpieza
            df234_T = pd.DataFrame(data234)
            df234 = df234_T.drop_duplicates() # Eliminar duplicados
            df234 = df234.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df234.columns = df234.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df234['OTROS'].str.contains('CONSUMO') # Filtro de Consumo
            condSancion = df234['OTROS'].str.contains('SANCION') # Filtro de Sancion
            condMedidor = df234['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df234['OTROS'].str.contains('RECONEXION') # Filtro de Reconexion
            condGasto = df234['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df234['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df234['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df234['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df234['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df234['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df234['OTROS'].str.contains('Caja de medidor') # Filtro de Reemplazo de Medidor
            condConv = df234['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df234['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df234 = df234[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df234 = df234.filter(items = ['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'TD/TC' ,'FOLIO/VOUCHER','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df234['INGRESOS'] = df234['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df234['FECHA'] = df234['FECHA'].apply(convertir_fecha)
            df234['FECHA'] = pd.to_datetime(df234['FECHA'], format='%d/%m/%Y')
            df234['FECHA'] = df234['FECHA'].dt.strftime('%d/%m/%Y')
            df234['OTROS'] = df234['OTROS'].str.upper()

            #df234 = df234.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df234 = df234.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df234 = df234[['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'FOLIO/VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df234.insert(0,'CUENTA', "BANREG 0234")

            df234.rename(columns={'EGRESOS':'RETIROS','FOLIO/VOUCHER':'FOLIO / VOUCHER'}, inplace=True)

            #print(df234)

            ###############  1351  ############### 

            sheet1351 = spreadsheet.worksheet("1351")

            # Obtener datos
            data1351 = sheet1351.get_all_records()

            # DataFrame y Limpieza
            df1351_T = pd.DataFrame(data1351)
            df1351 = df1351_T.drop_duplicates() # Eliminar duplicados
            df1351 = df1351.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df1351.columns = df1351.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df1351['OTROS'].str.contains('CONSUMO') # Filtro de Consumo
            condSancion = df1351['OTROS'].str.contains('SANCION') # Filtro de Sancion
            condMedidor = df1351['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df1351['OTROS'].str.contains('RECONEXION') # Filtro de Reconexion
            condGasto = df1351['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df1351['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df1351['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df1351['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df1351['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df1351['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df1351['OTROS'].str.contains('Caja de medidor') # Filtro de Reemplazo de Medidor
            condConv = df1351['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df1351['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df1351 = df1351[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df1351 = df1351.filter(items = ['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'TD/TC' ,'FOLIO/VOUCHER','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df1351['INGRESOS'] = df1351['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df1351['FECHA'] = df1351['FECHA'].apply(convertir_fecha)
            df1351['FECHA'] = pd.to_datetime(df1351['FECHA'], format='%d/%m/%Y')
            df1351['FECHA'] = df1351['FECHA'].dt.strftime('%d/%m/%Y')
            df1351['TD/TC'].replace({'Credito': 'TC', 'Debito': 'TD'}, inplace=True)
            df1351['OTROS'] = df1351['OTROS'].str.upper()
            #df1351 = df1351.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df1351 = df1351.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df1351 = df1351[['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'FOLIO/VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df1351.insert(0,'CUENTA', "BAJIO 1351")

            df1351.rename(columns={'EGRESOS':'RETIROS','FOLIO/VOUCHER':'FOLIO / VOUCHER'}, inplace=True)

            #print(df1351)

            ###############  7573  ############### 

            sheet7573 = spreadsheet.worksheet("7573")

            # Obtener datos
            data7573 = sheet7573.get_all_records()

            # DataFrame y Limpieza
            df7573_T = pd.DataFrame(data7573)
            df7573 = df7573_T.drop_duplicates() # Eliminar duplicados
            df7573 = df7573.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df7573.columns = df7573.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df7573['OTROS'].str.contains('CONSUMO') # Filtro de Consumo
            condSancion = df7573['OTROS'].str.contains('SANCION') # Filtro de Sancion
            condMedidor = df7573['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df7573['OTROS'].str.contains('RECONEXION') # Filtro de Reconexion
            condGasto = df7573['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df7573['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df7573['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df7573['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df7573['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df7573['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df7573['OTROS'].str.contains('Caja de medidor') # Filtro de Reemplazo de Medidor
            condConv = df7573['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df7573['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df7573 = df7573[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df7573 = df7573.filter(items = ['FECHA','DESCRIPCION','VOUCHERS','RETIROS', 'INGRESOS' ,'LOTE','OTROS','RECIBO']) # Eliminar columnas

            df7573['TD/TC'] = '' # Agregar columna TD/TC

            df7573['INGRESOS'] = df7573['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df7573['FECHA'] = pd.to_datetime(df7573['FECHA'], format='%d/%m/%Y')
            df7573['FECHA'] = df7573['FECHA'].dt.strftime('%d/%m/%Y')
            df7573['OTROS'] = df7573['OTROS'].str.upper()

            #df7573 = df7573.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df7573 = df7573.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df7573 = df7573[['FECHA','DESCRIPCION','RETIROS','INGRESOS', 'VOUCHERS','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df7573.insert(0,'CUENTA', "BBVA 7573")

            df7573.rename(columns={'VOUCHERS':'FOLIO / VOUCHER'}, inplace=True)

            #print(df7573)

            ###############  146  ############### 

            sheet146 = spreadsheet.worksheet("146")

            # Obtener datos
            data146 = sheet146.get_all_records()

            # DataFrame y Limpieza
            df146_T = pd.DataFrame(data146)
            df146 = df146_T.drop_duplicates() # Eliminar duplicados
            df146 = df146.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df146.columns = df146.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df146['OTROS'].str.contains('CONSUMO') # Filtro de Consumo
            condSancion = df146['OTROS'].str.contains('SANCION') # Filtro de Sancion
            condMedidor = df146['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df146['OTROS'].str.contains('RECONEXION') # Filtro de Reconexion
            condGasto = df146['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df146['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df146['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df146['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df146['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df146['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df146['OTROS'].str.contains('Caja de medidor') # Filtro de Reemplazo de Medidor
            condConv = df146['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df146['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df146 = df146[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe
            df146 = df146.filter(items = ['FECHA','DESCRIPCION','EGRESOS', 'INGRESOS' ,'TD/TC','FOLIO/VOUCHER','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df146['INGRESOS'] = df146['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df146['FECHA'] = pd.to_datetime(df146['FECHA'], format='%d/%m/%Y')
            df146['FECHA'] = df146['FECHA'].dt.strftime('%d/%m/%Y')
            df146['OTROS'] = df146['OTROS'].str.upper()

            #df146 = df146.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df146 = df146.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df146 = df146[['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'FOLIO/VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df146.insert(0,'CUENTA', "BAJIO 0146")

            df146.rename(columns={'EGRESOS':'RETIROS','FOLIO/VOUCHER':'FOLIO / VOUCHER'}, inplace=True)

            #print(df146)

            ###############  285  ############### 

            sheet285 = spreadsheet.worksheet("285")

            # Obtener datos
            data285 = sheet285.get_all_records()

            # DataFrame y Limpieza
            df285_T = pd.DataFrame(data285)
            df285 = df285_T.drop_duplicates() # Eliminar duplicados
            df285 = df285.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            df285.columns = df285.columns.str.strip() # Eliminar espacios de encabezados
            condConsumo = df285['OTROS'].str.contains('CONSUMO') # Filtro de Consumo
            condSancion = df285['OTROS'].str.contains('SANCION') # Filtro de Sancion
            condMedidor = df285['OTROS'].str.contains('MEDIDOR') # Filtro de Medidor
            condReconexion = df285['OTROS'].str.contains('RECONEXION') # Filtro de Reconexion
            condGasto = df285['OTROS'].str.contains('Gastos administrativos por pago extemporáneo') # Filtro de Gasto
            condTortuga = df285['OTROS'].str.contains('Pago tortuga') # Filtro de Pago Tortuga
            condReeAntena = df285['OTROS'].str.contains('Reemplazo de antena') # Filtro de Reemplazo de antena
            condAntMed = df285['OTROS'].str.contains('ANTENA MEDIDOR') # Filtro de Antena Medidor
            condHidra = df285['OTROS'].str.contains('Remosicon de estructura hidraulica') # Filtro de Estructura Hidraulica
            condReeMed = df285['OTROS'].str.contains('Reemplazo de medidor') # Filtro de Reemplazo de Medidor
            condCaja = df285['OTROS'].str.contains('Caja de medidor') # Filtro de Reemplazo de Medidor
            condConv = df285['OTROS'].str.contains('CONVENIO') # Filtro de Convenio
            #condREGEX = df285['LOTE'].str.match(r'^[A-Za-z0-9]{6}-[A-Za-z0-9]{3}O-[A-Za-z0-9]{3}$') # Filtro de REGEX

            df285 = df285[condConsumo | condSancion | condMedidor | condReconexion | condGasto | condTortuga | condReeAntena | condAntMed | condHidra | condReeMed | condCaja | condConv] # Filtrar Dataframe

            df285 = df285.filter(items = ['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'TD/TC' ,'FOLIO/VOUCHER','LOTE','OTROS','RECIBO']) # Eliminar columnas

            df285['INGRESOS'] = df285['INGRESOS'].replace({r'\$': '', r',': '', r'\.': ''}, regex=True).astype(float) / 100 # Ingresos a valores numéricos
            df285['FECHA'] = df285['FECHA'].apply(convertir_fecha)
            df285['FECHA'] = pd.to_datetime(df285['FECHA'], format='%d/%m/%Y')
            df285['FECHA'] = df285['FECHA'].dt.strftime('%d/%m/%Y')
            df285['OTROS'] = df285['OTROS'].str.upper()

            #df285 = df285.groupby(['FECHA','DESCRIPCION','RETIROS','FOLIO / VOUCHER','TD/TC','LOTE','OTROS','RECIBO'], as_index=False)['INGRESOS'].sum()

            df285 = df285.sort_values(by=['LOTE', 'FECHA']) # Ordenar elementos del Dataframe

            df285 = df285[['FECHA','DESCRIPCION','EGRESOS','INGRESOS', 'FOLIO/VOUCHER','TD/TC','LOTE','OTROS','RECIBO']] # Ordenar columnas del Dataframe

            df285.insert(0,'CUENTA', "BANREG 0285")

            df285.rename(columns={'EGRESOS':'RETIROS','FOLIO/VOUCHER':'FOLIO / VOUCHER'}, inplace=True)

            #print(df285)

            # Unir Dataframes
            dfBancos = pd.concat([df2239,df2162,df9721,df234,df1351,df7573,df146,df285], ignore_index=True)
            dfBancos['FECHA'] = pd.to_datetime(dfBancos['FECHA'], format='%d/%m/%Y')            
            df_MEDIDORES = dfBancos[dfBancos['OTROS'].str.startswith('MEDIDOR')]
            dfBancos = dfBancos[~dfBancos['OTROS'].str.startswith('MEDIDOR')]
            df_ENTREGAS = dfBancos[dfBancos['DESCRIPCION'].str.startswith('Entrega')]
            dfBancos = dfBancos[~dfBancos['DESCRIPCION'].str.startswith('Entrega')]
            df_MESANT = dfBancos[dfBancos['FECHA'].dt.month == 4]
            dfBancos = dfBancos[dfBancos['FECHA'].dt.month != 4]

            # Exportar a Excel
            fecha_hoy = datetime.now().strftime('%Y-%m-%d')
            nombre_archivo = f'BANCOS {fecha_hoy}.xlsx'
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
                dfBancos.to_excel(writer, sheet_name="CONSUMOS", index=False)
                df_MEDIDORES.to_excel(writer, sheet_name="MEDIDORES", index=False)
                df_ENTREGAS.to_excel(writer, sheet_name="ENTREGAS", index=False)
                df_MESANT.to_excel(writer, sheet_name="MES ANTERIOR", index=False)
                df1351_T.to_excel(writer, sheet_name="1351",index=False)
                df146_T.to_excel(writer, sheet_name="146",index=False)
                df234_T.to_excel(writer, sheet_name="234",index=False)
                df285_T.to_excel(writer, sheet_name="285",index=False)
                df7573_T.to_excel(writer, sheet_name="7573",index=False)
                df9721_T.to_excel(writer, sheet_name="9721",index=False)
                df2239_T.to_excel(writer, sheet_name="2239",index=False)
                df2162_T.to_excel(writer, sheet_name="2162",index=False)

            excepciones = ["CPSQRO-TOMO-000", "SNDQRO-TOMO-000", "CMLJAL-ABDO-0CC","CMLJAL-ABEO-0CC","CMLJAL-ACAO-0CC","CMLJAL-AGAO-0CC","CMLJAL-ALAO-0CC","CMLJAL-ALMO-0CC","CMLJAL-ALTO-0CC","CMLJAL-AMAO-0CC","CMLJAL-ARAO-0AA","CMLJAL-ARBO-0AA","CMLJAL-AREO-0CC","CMLJAL-ARRO-0CC","CMLJAL-AVEO-0CC","CMLJAL-AZAO-0CC","CMLJAL-BAMO-0CC","CMLJAL-BANO-0CC","CMLJAL-BIZO-0CC","CMLJAL-BOSO-0AA","CMLJAL-BUGO-0CC","CMLJAL-CAMO-0CC","CMLJAL-CAOO-0CC","CMLJAL-CARO-0CC","CMLJAL-CASO-0CC","CMLJAL-CEDO-0CC","CMLJAL-CEIO-0AA","CMLJAL-CEIO-0CC","CMLJAL-CIPO-0CC","CMLJAL-CIRO-0CC","CMLJAL-CYCO-0CC","CMLJAL-MACO-001","CMLJAL-MACO-002","CMLJAL-DURO-0CC","CMLJAL-EBAO-0CC","CMLJAL-ENEO-0CC","CMLJAL-ENRO-0AA","CMLJAL-EUCO-0CC","CMLJAL-FLAO-0CC","CMLJAL-FREO-0CC","CMLJAL-GARO-0CC","CMLJAL-GERO-0CC","CMLJAL-GOTO-000","CMLJAL-GRAO-0CC","CMLJAL-HORO-0AA","CMLJAL-HUIO-0CC","CMLJAL-JACO-0CC","CMLJAL-JARO-0AA","CMLJAL-LATO-0CC","CMLJAL-LAUO-0CC","CMLJAL-LOMO-0AA","CMLJAL-MAGO-0CC","CMLJAL-MALO-0CC","CMLJAL-MANO-069","CMLJAL-MANO-0CC","CMLJAL-MAPO-0CC","CMLJAL-MEZO-0CC","CMLJAL-NARO-0CC","CMLJAL-NOGO-0CC","CMLJAL-OLIO-0CC","CMLJAL-OLMO-0AA","CMLJAL-OLMO-0CC","CMLJAL-OYAO-0CC","CMLJAL-PLUO-001","CMLJAL-PLUO-002","CMLJAL-PRAO-0AA","CMLJAL-ROBO-0AA","CMLJAL-ROBO-0CC","CMLJAL-SABO-0CC","CMLJAL-SAUO-0CC","CMLJAL-SEQO-0AA","CMLJAL-SEQO-0CC","CMLJAL-TUNO-0CC","CMLJAL-VALO-0AA","CMLJAL-YUCO-0CC","CDMSLP-ABDO-0CC","CDMSLP-ABEO-0CC", "CDMSLP-ACAO-0CC", "CDMSLP-AGAO-0CC","CDMSLP-AHUO-0CC","CDMSLP-ALAO-0CC", "CDMSLP-ALEO-0CC", "CDMSLP-ALIO-0CC","CDMSLP-ALMO-0CC","CDMSLP-ALTO-0CC","CDMSLP-AMAO-0CC","CDMSLP-ARAO-0AA","CDMSLP-ARBO-0AA","CDMSLP-AREO-0CC","CDMSLP-ARRO-0CC","CDMSLP-AVEO-0CC","CDMSLP-AZAO-0CC","CDMSLP-BAMO-0CC","CDMSLP-BANO-0CC","CDMSLP-BIZO-0CC","CDMSLP-BOSO-0AA","CDMSLP-BUGO-0CC","CDMSLP-CACO-0AA","CDMSLP-CAOO-0CC","CDMSLP-CARO-0CC","CDMSLP-CRRO-0CC","CDMSLP-CASO-CF 14.1","CDMSLP-CASO-CC-01","CDMSLP-CASO-0CC","CDMSLP-CASO-0CC","CDMSLP-CEDO-0CC","CDMSLP-CEIO-0CC","CDMSLP-CERO-0CC","CDMSLP-CIPO-0CC","CDMSLP-CIRO-0CC","CDMSLP-COLO-0AA","CDMSLP-CONO-0AA","CDMSLP-CUMO-0AA","CDMSLP-CYCO-0CC","CDMSLP-DUNO-0AA","CDMSLP-DURO-0CC","CDMSLP-EBAO-0CC","CDMSLP-ENEO-0CC","CDMSLP-ENRO-0AA","CDMSLP-EUCO-0CC","CDMSLP-FLAO-0CC","CDMSLP-FLOO-0AA","CDMSLP-FREO-0CC","CDMSLP-GARO-0CC","CDMSLP-GERO-0CC","CDMSLP-GRAO-0CC","CDMSLP-HORO-0AA","CDMSLP-HUIO-0CC","CDMSLP-JACO-0CC","CDMSLP-JARO-AV-03","CDMSLP-JARO-0AA","CDMSLP-JUNO-0CC","CDMSLP-KAUO-0CC","CDMSLP-LAMO-0CC","CDMSLP-LATO-0CC","CDMSLP-LAUO-0CC","CDMSLP-LILO-0CC","CDMSLP-LOMO-0AA","CDMSLP-MAGO-0CC","CDMSLP-MALO-0CC","CDMSLP-MANO-0CC","CDMSLP-MAPO-0CC","CDMSLP-MEZO-0CC","CDMSLP-NARO-0CC","CDMSLP-NOGO-0CC","CDMSLP-OLMO-0CC","CDMSLP-PALO-0AA","CDMSLP-PEDO-0AA","CDMSLP-PINO-039","CDMSLP-PINO-0CC","CDMSLP-PIRO-0CC","CDMSLP-PRAO-0AA","CDMSLP-RIBO-0AA","CDMSLP-ROBO-0CC","CDMSLP-ROMO-0CC","CDMSLP-SABO-0CC","CDMSLP-SAUO-0CC","CDMSLP-TABO-0CC","CDMSLP-TULO-0CC","CDMSLP-TUNO-0CC","CDMSLP-VALO-0AA","CDMSLP-VALO-0CC","CDMSLP-YUCO-0CC","CDMSLP-ZARO-0CC","CS1GTO-ABEO-0CC","CS1GTO-AVEO-0CC","CS1GTO-CASO-0CC","CS1GTO-CEDO-0CC","CS1GTO-CEIO-0CC","CS1GTO-ENCO-0CC","CS1GTO-MAPO-0CC","CS1GTO-NOGO-0CC","CS1GTO-OYAO-0CC","CS1GTO-ROBO-0CC","CS1GTO-S1O-0AA","CS1GTO-SEQO-0CC","CS2GTO-ABEO-0CC","CS2GTO-ACAO-0CC","CS2GTO-AGAO-0CC","CS2GTO-ALMO-0CC","CS2GTO-AMAO-0CC","CS2GTO-ARCO-0CC","CS2GTO-ARCO-MACRO ACCESO","CS2GTO-BAMO-0CC","CS2GTO-BIZO-0CC","CS2GTO-CAMO-0CC","CS2GTO-CAOO-0CC","CS2GTO-CRRO-0CC","CS2GTO-CEDO-0CC","CS2GTO-CERO-0CC","CS2GTO-CIRO-0CC","CS2GTO-DURO-0CC","CS2GTO-ENEO-0CC","CS2GTO-FREO-0CC","CS2GTO-GERO-0CC","CS2GTO-GRAO-0CC","CS2GTO-HORO-0AA","CS2GTO-HUIO-0CC","CS2GTO-LATO-0CC","CS2GTO-MAGO-0CC","CS2GTO-MANO-0CC","CS2GTO-MEZO-0CC","CS2GTO-NARO-0CC","CS2GTO-NOGO-0CC","CS2GTO-PINO-0CC","CS2GTO-ROBO-0CC","CS2GTO-ROMO-0CC","CS2GTO-SABO-0CC","CS2GTO-TULO-0CC","CS2GTO-YUCO-0CC","CS2GTO-ZARO-0CC","CMCQRO-ARCO-0AA","CMCQRO-AHUO-A-0CC","CMCQRO-AHUO-B-0CC","CMCQRO-ALEO-0CC","CMCQRO-ALTO-0CC","CMCQRO-AREO-A-0CC","CMCQRO-AREO-B-0CC","CMCQRO-BANO-0CC","CMCQRO-CARO-A-0CC","CMCQRO-CARO-B-0CC","CPCQRO-CUPO-0CC","CMCQRO-EUCO-A-0CC","CMCQRO-EUCO-B-0CC","CMCQRO-MAPO-0CC","CPCQRO-OPUO-0CC","CMCQRO-PIRO-0CC","CMCQRO-SEQO-0CC","CDMQRO-ABDO-0CC","CDMQRO-ABEO-0CC","CDMQRO-ACAO-0CC","CDMQRO-ALMO-0CC","CDMQRO-AMAO-0AA","CDMQRO-AMAO-0CC","CDMQRO-ARCO-0CC","CDMQRO-AVEO-0CC","CDMQRO-BUGO-0CC","CDMQRO-CAOO-086","CDMQRO-CAOO-0CC","CDMQRO-ACEO-CISTERNA","CDMQRO-CAAO-C75","CDMQRO-CAPO-C","CDMQRO-CASO-0CC","CDMQRO-CEDO-0CC","CDMQRO-CEIO-0CC","CDMQRO-CEIO-0AA","CDMQRO-CERO-0CC","CDMQRO-CIPO-0CC","CDMQRO-COLO-0AA","CDMQRO-ENCO-0CC","CDMQRO-EUCO-0CC","CDMQRO-FREO-0CC","CDMQRO-GRAO-0CC","CDMQRO-JACO-0CC","CDMQRO-LAUO-0CC","CDMQRO-MAPO-0CC","CDMQRO-NARO-0CC","CDMQRO-NOGO-074","CDMQRO-NOGO-0CC","CDMQRO-NALO-0CC","CDMQRO-OLIO-0CC","CDMQRO-OLMO-0CC","CDMQRO-OYAO-0CC","CDMQRO-ROBO-0CC","CDMQRO-SEQO-0CC","CDMQRO-ARBO-0AA","CDMQRO-BOSO-0AA","CDMQRO-JARO-0AA","CDMQRO-INDO-0AA","CDMQRO-LOMO-0AA","CDMQRO-FLOO-0AA","CDMQRO-VALO-0AA","CDMQRO-GRAO-0AA","CDMQRO-JACO-0AA","CDMQRO-SEQO-0AA","CDMQRO-TOMO-000","CMMQRO-ACOO-0CC","CMMQRO-AJUO-0CC","CMMQRO-ASPO-0CC","CMMQRO-CRVO-0CC","CMMQRO-CIVO-0CC","CMMQRO-EVEO-0CC","CMMQRO-FUJO-0AA","CMMQRO-FUJO-0CC","CMMQRO-HUAO-0CC","CMMQRO-MAKO-0CC","CMMQRO-MANO-0CC","CMMQRO-MARO-0CC","CMMQRO-MBAO-0CC","CMMQRO-MONO-0CC","CMMQRO-NCEO-0CC","CMMQRO-PALO-0CC","CMMQRO-PIRO-0CC","CMMQRO-POPO-0CC","CMMQRO-PUMO-0CC","CMMQRO-QUIO-0CC","CMMQRO-TABO-0CC","CMMQRO-ALPO-0AA","CMMQRO-ENRO-0AA","CMMQRO-HIMO-0AA","CMMQRO-ANDO-0AA","CMMQRO-VOLO-0AA","CMMQRO-PALO-0AA","CMMQRO-PIRO-0AA","CMMQRO-VILO-0CC","CMMQRO-ZAMO-0CC"]
            simulados = pd.DataFrame({
                "LOTE": excepciones,
                "FECHA": ["01/01/2000"] * len(excepciones)
            })
            dfBancos = pd.concat([dfBancos, simulados], ignore_index=True)
            dfBancos['FECHA'] = pd.to_datetime(dfBancos['FECHA'], format='%d/%m/%Y') # Formato de Fecha
            dfBancos['FOLIO / VOUCHER'].fillna('', inplace=True)
            fecha_corte = pd.to_datetime(fecha_corte, format='%d/%m/%Y')
            dfCobros = pd.read_excel(uploaded_file, sheet_name="COBRO 2025", header=1, usecols="F:GR") # Tabla de Cobro
            dfCobros = dfCobros[dfCobros.iloc[:, 0].notna()] # Valores no nulos en primer columna
            dfCobros = dfCobros.applymap(lambda x: x.strip() if isinstance(x, str) else x) # Eliminar espacios
            dfCobros = dfCobros.loc[:, ['CUENTA POR COBRAR MAY','NOMENCLATURA OOAM','PAGO A TIEMPO MAY', 'PAGO VENCIDO MAY','PAGO ATRASADO MAY','FECHA DE PAGO MAY','FACT/VOUCHER MAY','BANCO MAY']] # Eliminar columnas
            dfCobros['FECHA DE PAGO MAY'] = dfCobros['FECHA DE PAGO MAY'].astype(str) # Pasar a string la fecha
            # Inicializar
            dfCobros['PAGO A TIEMPO MAY'] = 0
            dfCobros['PAGO VENCIDO MAY'] = 0
            dfCobros['PAGO ATRASADO MAY'] = 0
            dfCobros['FECHA DE PAGO MAY'] = ""
            dfCobros['FACT/VOUCHER MAY'] = ""
            dfCobros['BANCO MAY'] = ""
            # Iterar
            for i, row_bancos in dfBancos.iterrows():
                mask = dfCobros['NOMENCLATURA OOAM'] == row_bancos['LOTE']  # Coincidencia
                for j, row_cobro in dfCobros[mask].iterrows():
                    # Verificar si NOMENCLATURA está en la lista de excepciones
                    if row_cobro['NOMENCLATURA OOAM'] in excepciones:
                        dfCobros.at[j, 'PAGO A TIEMPO MAY'] = row_cobro['CUENTA POR COBRAR MAY']
                        dfCobros.at[j, 'FECHA DE PAGO MAY'] = fecha_corte.strftime('%d/%m/%y')
                        dfCobros.at[j, 'BANCO MAY'] = "PAGO SIMULADO"
                        continue
                    # Convertir fecha a cadena y obtener el mes del pago
                    fecha_str = row_bancos['FECHA'].strftime('%d/%m/%y')
                    mes_pago = row_bancos['FECHA'].month  # Obtener el mes del pago
                    # Condición para pagos en abril
                    if mes_pago == 4:
                        dfCobros.at[j, 'PAGO ATRASADO MAY'] += row_bancos['INGRESOS']
                        if dfCobros.at[j, 'FECHA DE PAGO MAY']:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] += f", {fecha_str}"
                        else:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] = fecha_str
                    elif row_bancos['FECHA'] <= fecha_corte:
                        dfCobros.at[j, 'PAGO A TIEMPO MAY'] += row_bancos['INGRESOS']
                        if fecha_str not in dfCobros.at[j, 'FECHA DE PAGO MAY']:
                            if dfCobros.at[j, 'FECHA DE PAGO MAY']:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] += f", {fecha_str}"
                            else:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] = fecha_str
                    else:
                        dfCobros.at[j, 'PAGO VENCIDO MAY'] += row_bancos['INGRESOS']
                        if fecha_str not in dfCobros.at[j, 'FECHA DE PAGO MAY']:
                            if dfCobros.at[j, 'FECHA DE PAGO MAY']:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] += f", {fecha_str}"
                            else:
                                dfCobros.at[j, 'FECHA DE PAGO MAY'] = fecha_str

                    # Voucher y Cuenta
                    folio_voucher = str(row_bancos['FOLIO / VOUCHER'])
                    if folio_voucher not in dfCobros.at[j, 'FACT/VOUCHER MAY']:
                        if dfCobros.at[j, 'FACT/VOUCHER MAY']:
                            dfCobros.at[j, 'FACT/VOUCHER MAY'] += f", {folio_voucher}"
                        else:
                            dfCobros.at[j, 'FACT/VOUCHER MAY'] = folio_voucher
                    cuenta = str(row_bancos['CUENTA'])
                    if cuenta not in dfCobros.at[j, 'BANCO MAY']:
                        if dfCobros.at[j, 'BANCO MAY']:
                            dfCobros.at[j, 'BANCO MAY'] += f", {cuenta}"
                        else:
                            dfCobros.at[j, 'BANCO MAY'] = cuenta

            dfCobros['FECHA DE PAGO MAY'] = dfCobros['FECHA DE PAGO MAY'].str.lstrip(', ')
            dfCobros['FACT/VOUCHER MAY'] = dfCobros['FACT/VOUCHER MAY'].str.lstrip(', ')
            dfCobros['BANCO MAY'] = dfCobros['BANCO MAY'].str.lstrip(', ')
            dfCobros = dfCobros.loc[:, ['PAGO A TIEMPO MAY', 'PAGO VENCIDO MAY','PAGO ATRASADO MAY','FECHA DE PAGO MAY','FACT/VOUCHER MAY','BANCO MAY']]
            dfCobros_1 = dfCobros.loc[:, ['PAGO A TIEMPO MAY']]
            dfCobros_1_1 = dfCobros.loc[:, ['PAGO VENCIDO MAY']]
            dfCobros_2 = dfCobros.loc[:, ['FECHA DE PAGO MAY','FACT/VOUCHER MAY','BANCO MAY']]
            wbCobro = load_workbook(uploaded_file)
            sheetCobro = wbCobro['COBRO 2025']

            for r_idx, row in dfCobros_1.iterrows():
                value = row.iloc[0]
                if value != 0 : 
                    sheetCobro.cell(row=r_idx + 3, column=185, value=value)  # Ajustar ubicación en Excel

            for r_idx, row in dfCobros_1_1.iterrows():
                value = row.iloc[0]
                if value != 0 : 
                    sheetCobro.cell(row=r_idx + 3, column=186, value=value)  # Ajustar ubicación en Excel                   

            for r_idx, row in dfCobros_2.iterrows():
                value_a_tiempo = sheetCobro.cell(row=r_idx + 3, column=185).value
                value_vencido = sheetCobro.cell(row=r_idx + 3, column=186).value
                if (value_a_tiempo is not None and value_a_tiempo !="") or (value_vencido is not None and value_vencido !=""):   
                     for c_idx, value in enumerate(row):
                        sheetCobro.cell(row=r_idx + 3, column=c_idx + 188, value=value)  # Ajustar ubicación en Excel

            output_filename = "Tabla_Actualizada.xlsx"
            wbCobro.save(output_filename)

            zipfilename = "Tabla_Bancos_Actualizados.zip"
            with zipfile.ZipFile(zipfilename,"w") as zipf:
                zipf.write(nombre_archivo)
                zipf.write(output_filename)

            msjcobro_proceso.write("Proceso completado")
            with open(zipfilename, "rb") as file:
                st.download_button(
                    label="Descargar bancos y tabla actualizada",
                    data=file,
                    file_name=zipfilename,
                    mime="application/zip"
                    )

elif opcion == "Lecturas":
    st.title("Actualización de lecturas - ABRIL")
    uploaded_csv = st.file_uploader("Elige tu archivo CSV (Solo Temetra)", type=["csv"])
    uploaded_excel = st.file_uploader("Selecciona tu tabla de cobro", type=["xlsx", "xls"])

    if uploaded_csv and uploaded_excel:
        if st.button("Procesar"):
            msjlec_proceso = st.empty()
            msjlec_proceso.write("Procesando archivo...")
            dfLec = pd.read_csv(uploaded_csv)
            dfLec['Post Code'] = dfLec['Post Code'].str.upper()
            excepciones = ["AVINON ACCESO","HACIENDA SL", "INDUSTRIAL MADERAS 3000","PTAR","COMERCIAL C270","E07GRA","GENERAL MACRO", "ARCO DE ACCESO AA", "ARCO GENERAL SL", "ARCO DE ACCESO 0AA", "MACRO ARCO CASETA ENTRADA", "CLUB LEON MACRO PTAR", "MACRO ARCO 2", "POLIGONO CLUB LEON MACRO", "SOMNUS","UNKNOWN", "109ET11JARAA", "109ET10ENRAA","109ET17RIBAA","109ET09ARAAA","BODEGA CONSTRUCCION MACRO", "VITIVINICOLA 10 Y 11", "CASETA POLICIAL CC-01","CASA DE VENTAS CF 14.1","JARDIN MADERAS AV-03", "109ET18VALAA", "PROVENZA MACROMEDIDOR GENERAL", "C089", "TARAY GENERAL","COMERCIAL C083-2","C083", "COMERCIAL 018","CMQ V38-2","COMERCIAL C212","AMATE AA","C GRAL-C","D GRAL-D","A GRAL-A","B GRAL-B","COMERCIAL C27"]
            dfExcepciones = dfLec[dfLec['Post Code'].isin(excepciones)]

            dfLec = dfLec[~dfLec['Post Code'].isin(excepciones)]
            dfLec = dfLec.loc[:, ["Route", "Post Code", "Index", "Account Address"]]
            dfLec['Post Code'] = dfLec['Post Code'].replace('', None)
            dfLec['Post Code'] = dfLec['Post Code'].fillna(dfLec['Account Address'])
            dfLec['Post Code'] = dfLec['Post Code'].str.replace(',', '')
            dfLec['Post Code'] = dfLec['Post Code'].str.replace('/', ' ')

            regexPRV = r'\b(PROVENZA)\b'
            dfPRV = dfLec[dfLec['Post Code'].str.contains(regexPRV, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexPRV, regex=True)]

            regexPremium = r'\b(PREMIUM)\b'
            dfPremium = dfLec[dfLec['Post Code'].str.contains(regexPremium, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexPremium, regex=True)]

            regexUC = r'\b(UC)\b'
            dfUC = dfLec[dfLec['Post Code'].str.contains(regexUC, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexUC, regex=True)]

            regexAA = r'\b(AA|0AA)\b'
            dfAA = dfLec[dfLec['Post Code'].str.contains(regexAA, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexAA, regex=True)]

            regexCC = r'\b(CC|0CC|PALAPA)\b'
            dfCC = dfLec[dfLec['Post Code'].str.contains(regexCC, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexCC, regex=True)]

            regexMacro = r'\b(MACRO|MAC)\b'
            dfMacro = dfLec[dfLec['Post Code'].str.contains(regexMacro, regex=True)]
            dfLec = dfLec[~dfLec['Post Code'].str.contains(regexMacro, regex=True)]

            def separar(cadena):
                if pd.isna(cadena):
                    return pd.Series(['', ''])
                cadena = str(cadena)
                strings = ' '.join(re.findall('[a-zA-Z]+', cadena))
                numeros = ' '.join(re.findall('\d+', cadena))
                return pd.Series([strings, numeros])

            def separar2(cadena):
                if pd.isna(cadena):
                        return pd.Series(['', ''])
                cadena = str(cadena)
                match = re.match(r'([a-zA-Z ]+\d*) (\d+)', cadena)
                if match:
                    return pd.Series([match.group(1).strip(), match.group(2).strip()])
                else:
                    return pd.Series([cadena, ''])

            exc = {
                'ABEDUL': 'ABDO',
                'ACANTO': 'ACNO',
                'AHUEHUETE A': 'AHUO-A',
                'AHUEHUETE B': 'AHUO-B',
                'ALESIA': 'ALSO',
                'ALIANTO': 'ALNO',
                'ALPAMAYO': 'APAO',
                'AMAZONAS': 'AMZO',
                'ARAUCA': 'ARUO',
                'ARECA A': 'AREO-A',
                'ARECA B': 'AREO-B',
                'AZZURRO': 'AZUO',
                'CARANDAY A': 'CARO-A',
                'CARANDAY B': 'CARO-B',
                'CARDON': 'CRDO',
                'CARRIZO': 'CRRO',
                'CERVINO': 'CRVO',
                'CIPERO': 'CIEO',
                'CRETA': 'CRTO',
                'DURIAN': 'DUIO',
                'GARDENIA': 'GADO',
                'GARONA': 'GAOO',
                'GRAN CAÑON': 'GCAO',
                'GRAN SENDA': 'GSEO',
                'INCUDINE': 'INUO',
                'INDUSTRIAL CIUDAD MADERAS': 'COMO',
                'JUNIPERO': 'JUIO',
                'LA LOMA': 'LOMO',
                'LOS ANDES': 'ANDO',
                'LOS VOLCANES': 'VOLO',
                'MANASLU': 'MASO',
                'MARATTIA': 'MRAO',
                'MARSELLA': 'MRSO',
                'MARSILEA': 'MRSO',
                'MELIA': 'MEIO',
                'MONSERRATE': 'MOSO',
                'MONTE BALDO': 'MBAO',
                'MONTE BLANC': 'MBLO',
                'MONTE ORO': 'MORO',
                'MONTE DE ORO': 'MORO',
                'MONTE VERDE': 'MVEO',
                'MONVISO': 'MOVO',
                'MONTEVIDEO': 'MOTO',
                'NARCISO': 'NACO',
                'NARDO': 'NADO',
                'NARMADA': 'NAMO',
                'NUEVA DELHI': 'NDEO',
                'NUEVO ALAMO': 'NALO',
                'NUEVO CASTAÑO': 'NCAO',
                'NUEVO CEIBA': 'NCEO',
                'PARAISO': 'PRSO',
                'PARANA': 'PRNO',
                'PENA BLANCA': 'PBLO',
                'PENABLANCA': 'PBLO',
                'PEÑA BLANCA': 'PBLO',
                'PIRINEOS': 'PIIO',
                'PREMIUM 1': 'PR1O',
                'PREMIUM 2': 'PR2O',
                'PREMIUM 3': 'PR3O',
                'PREMIUM 4': 'PR4O',
                'PREMIUM 5': 'PR5O',
                'PREMIUM 6': 'PR6O',
                'PREMIUM 7': 'PR7O',
                'PREMIUM 8': 'PR8O',
                'PREMIUM ARBOLEDAS': 'ARBO',
                'PREMIUM BOSQUES': 'BOSO',
                'PREMIUM COLINAS': 'COLO',
                'PREMIUM CONIFERAS': 'CONO',
                'PREMIUM ENRAMADA': 'ENRO',
                'PREMIUM FLORESTA 1': 'FL1O',
                'PREMIUM FLORESTA 2': 'FL2O',
                'PREMIUM HORTALIZAS': 'HORO',
                'PREMIUM JARDINES': 'JARO',
                'PREMIUM LOMAS': 'LOMO',
                'PREMIUM PALMAR': 'PAMO',
                'PREMIUM VALLE': 'VALO',
                'PREMIUM ZARZAL': 'ZAZO',
                'PUNTA CAVALLI': 'CAVO',
                'S1': 'S1O',
                'SAUCO': 'SACO',
                'VIU PANORAMA': 'VPNO'
            }

            def NOM(desarrollo, condominio, lote):
                condominio_mod = exc.get(condominio,condominio[:3]+'O')
                return f"{desarrollo}-{condominio_mod}-{str(lote).zfill(3)}"

            # Nomenclatura
            dfLec[['Condominio', 'Lote']] = dfLec['Post Code'].apply(separar)
            dfLec['Lote'] = dfLec['Lote'].astype(str).str.zfill(3)
            dfLec = dfLec.loc[:, ["Route", "Condominio", "Lote", "Index"]]
            dfLec = dfLec.rename(columns={'Route': 'Desarrollo', 'Index':'Lectura'})
            dfLec['Nomenclatura'] = dfLec.apply(lambda row: NOM(row['Desarrollo'], row['Condominio'], row['Lote']), axis=1)
            dfLec = dfLec.loc[:,['Nomenclatura', 'Lectura']]

            # Provenza
            if not dfPRV.empty:
                dfPRV['Post Code'] = dfPRV['Post Code'].str.replace('PROVENZA','',regex=False).str.strip()
                dfPRV['Lote'] = "MAC"
                dfPRV['Nomenclatura'] = dfPRV.apply(lambda row: NOM(row['Route'], row['Post Code'], row['Lote']), axis=1)
                dfPRV = dfPRV.loc[:,['Nomenclatura', 'Index']]
                dfPRV = dfPRV.rename(columns={'Index':'Lectura'})

            # Premium
            if not dfPremium.empty:
                dfPremium[['Condominio', 'Lote']] = dfPremium['Post Code'].apply(separar2)
                dfPremium['Lote'] = dfPremium['Lote'].astype(str).str.zfill(3)
                dfPremium['Nomenclatura'] = dfPremium.apply(lambda row: NOM(row['Route'], row['Condominio'], row['Lote']), axis=1)
                dfPremium = dfPremium.loc[:,['Nomenclatura', 'Index']]
                dfPremium = dfPremium.rename(columns={'Index':'Lectura'})

            # UC
            if not dfUC.empty:
                dfUC['Post Code'] = dfUC['Post Code'].str.replace('UC','',regex=False)
                dfUC['Post Code'] = dfUC['Post Code'].str.replace('0AA','',regex=False)
                dfUC['Post Code'] = dfUC['Post Code'].str.replace('AA','',regex=False).str.strip()
                dfUC['Lote'] = "0AA"
                dfUC['Nomenclatura'] = dfUC.apply(lambda row: NOM(row['Route'], row['Post Code'], row['Lote']), axis=1)
                dfUC = dfUC.loc[:,['Nomenclatura', 'Index']]
                dfUC = dfUC.rename(columns={'Index':'Lectura'})

            # AA
            if not dfAA.empty:
                dfAA['Post Code'] = dfAA['Post Code'].str.replace('0AA', '', regex=False)
                dfAA['Post Code'] = dfAA['Post Code'].str.replace('AA', '', regex=False)
                dfAA['Post Code'] = dfAA['Post Code'].str.strip()
                dfAA['Lote'] = "0AA"
                dfAA['Nomenclatura'] = dfAA.apply(lambda row: NOM(row['Route'], row['Post Code'], row['Lote']), axis=1)
                dfAA = dfAA.loc[:,['Nomenclatura', 'Index']]
                dfAA = dfAA.rename(columns={'Index':'Lectura'})

            # CC
            if not dfCC.empty:
                dfCC['Post Code'] = dfCC['Post Code'].str.replace('0CC', '', regex=False)
                dfCC['Post Code'] = dfCC['Post Code'].str.replace('CC', '', regex=False)
                dfCC['Post Code'] = dfCC['Post Code'].str.replace('PALAPA', '', regex=False)
                dfCC['Post Code'] = dfCC['Post Code'].str.strip()
                dfCC['Lote'] = "0CC"
                dfCC['Nomenclatura'] = dfCC.apply(lambda row: NOM(row['Route'], row['Post Code'], row['Lote']), axis=1)
                dfCC = dfCC.loc[:,['Nomenclatura', 'Index']]
                dfCC = dfCC.rename(columns={'Index':'Lectura'})

            # Macro
            if not dfMacro.empty:
                dfMacro['Post Code'] = dfMacro['Post Code'].str.replace('MACRO','',regex=False)
                dfMacro['Post Code'] = dfMacro['Post Code'].str.replace('MAC','',regex=False).str.strip()
                dfMacro['Lote'] = "MAC"
                dfMacro['Nomenclatura'] = dfMacro.apply(lambda row: NOM(row['Route'], row['Post Code'], row['Lote']), axis=1)
                dfMacro = dfMacro.loc[:,['Nomenclatura', 'Index']]
                dfMacro = dfMacro.rename(columns={'Index':'Lectura'})

            dfConcat = pd.concat([dfLec, dfPRV, dfPremium, dfUC, dfAA, dfCC, dfMacro])
            dfConcat = dfConcat.loc[:, ['Nomenclatura', 'Lectura']]

            cnc_name = "Concentrado.xlsx"
            with pd.ExcelWriter(cnc_name) as writer:
                dfConcat.to_excel(writer, sheet_name='LECTURAS', index=False)
                dfExcepciones.to_excel(writer, sheet_name='SIN IDENTIFICAR', index=False)

            dfLecturas = pd.read_excel(uploaded_excel, sheet_name="COBRO 2025", header=1, usecols="F:GR")  # Tabla de Cobro
            dfLecturas = dfLecturas[dfLecturas.iloc[:, 0].notna()]  # Filtrar valores no nulos en la primera columna
            dfLecturas = dfLecturas.applymap(lambda x: x.strip() if isinstance(x, str) else x)  # Eliminar espacios
            dfLecturas = dfLecturas.loc[:, ['NOMENCLATURA OOAM', 'LECTURA EMITIDA MAY']]  # Seleccionar columnas
            dfLecturas['LECTURA EMITIDA MAY'] = 0  # Inicializar columna
            dfConcat['Nomenclatura'] = dfConcat['Nomenclatura'].astype(str)
            dfLecturas['NOMENCLATURA OOAM'] = dfLecturas['NOMENCLATURA OOAM'].astype(str)
            
            for i, rowLecturas in dfLecturas.iterrows():
                for j, rowConcat in dfConcat.iterrows():
                    if rowLecturas['NOMENCLATURA OOAM'] == rowConcat['Nomenclatura']:
                        dfLecturas.at[i, 'LECTURA EMITIDA MAY'] = rowConcat['Lectura']
                        break

            wbLectura = load_workbook(uploaded_excel)
            sheetLectura = wbLectura['COBRO 2025']

            for r_idx, value in enumerate(dfLecturas['LECTURA EMITIDA MAY'], start=0):
                sheetLectura.cell(row=r_idx + 3, column=178, value=value)

            output_filename = "Tabla_Actualizada.xlsx"
            wbLectura.save(output_filename) 

            zipfilename = "Tabla_Lecturas_Actualizadas.zip"
            with zipfile.ZipFile(zipfilename,"w") as zipf:
                zipf.write(cnc_name)
                zipf.write(output_filename)

            msjlec_proceso.write("Proceso completado")
            with open(zipfilename, "rb") as file:
                st.download_button(
                    label="Descargar lecturas y tabla actualizada",
                    data=file,
                    file_name=zipfilename,
                    mime="application/zip"
                    )